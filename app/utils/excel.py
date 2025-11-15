from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ..models import Order, OrderItem


def generate_orders_excel(orders: Iterable[Order]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Header
    headers = [
        "Order ID",
        "Order Code",
        "Created At",
        "Customer",
        "Status",
        "Item Name",
        "Quantity",
        "Unit Price",
        "Line Total",
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    total_amount = Decimal("0.00")
    row_index = 1

    item_aggregates: dict[str, dict[str, Decimal | int]] = {}

    for order in orders:
        for item in order.items:
            row_index += 1
            ws.append(
                [
                    order.id,
                    order.order_code,
                    order.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                    order.customer_name,
                    order.status,
                    item.item_name,
                    item.quantity,
                    float(item.unit_price),
                    float(item.line_total),
                ]
            )
            total_amount += item.line_total

            # Aggregate by item_name
            agg = item_aggregates.setdefault(
                item.item_name,
                {"quantity": 0, "amount": Decimal("0.00")},
            )
            agg["quantity"] += item.quantity
            agg["amount"] += item.line_total

    # Totals section
    row_index += 2
    ws.cell(row=row_index, column=1, value="Total Amount:")
    ws.cell(row=row_index, column=2, value=float(total_amount))

    # Item aggregates section
    row_index += 2
    ws.cell(row=row_index, column=1, value="Item Summary")
    ws.cell(row=row_index, column=1).font = Font(bold=True)
    row_index += 1

    ws.append(["Item Name", "Total Quantity", "Total Amount"])
    for col in range(1, 4):
        cell = ws.cell(row=row_index, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for name, agg in item_aggregates.items():
        row_index += 1
        ws.append([name, agg["quantity"], float(agg["amount"])])

    # Auto-fit-ish column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max(12, length + 2)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
